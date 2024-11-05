"""--------------------------------------------------------"""
""" Parte que ler os  link da listagem                     """
"""--------------------------------------------------------"""
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import sys, os, random, time, datetime, traceback, requests_cache, logging
from openpyxl import Workbook, load_workbook  # Importe load_workbook
from funcoes_especificas import get_ceps_por_logradouro,extrair_numeros,encontrar_tipo_imovel,parse_address,remover_parte_texto,extrair_e_formatar_data,separar_precos
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from threading import Thread, Semaphore, Lock


requests_cache.install_cache('cache')

# Caminho da planilha
caminho_da_planilha = sys.argv[1]

# Caminho do webchorme 
caminho_do_webchrome = sys.argv[2]

#filtro pro cidade
caminho_do_filtro=sys.argv[3]

#tipo de filtro pode ser venda 
tipo_de_filtro=sys.argv[4]

#maximo de theads

maximo_theads=sys.argv[5]

tipo_processo=sys.argv[6] # gerarlink pra gerar arquivo texto lerlinks para gerar a ´pal

cidade_para_nome_arquivo=sys.argv[7]  # Aqui  vai o nome do arquivo que será gerado exe:  aquivo_niteroi.tct


# Configurando opções do Chrome
chrome_options = Options()
chrome_options.page_load_strategy="normal"
chrome_options.set_capability('goog:logginPrefs',{'performance': 'ALL'})
## Comente esta linha para ver o navegador
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--incognito")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-gpu")  # Desativa a aceleração de hardware
chrome_options.add_argument("--disable-extensions") 
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
#requests_cache.install_cache('cache')
# Inicializando o ChromeDriver
# Caminho para o ChromeDriver
chrome_driver_path = caminho_do_webchrome
#"C:\Users\User\ChromeWithDriver\chromedriver.exe"
# Inicializando um conjunto para armazenar links únicos
links_imoveis = set()
# Número da página inicial
pagina_atual = 1
cont = 0 
# PAgina por scrol de tela
pagina_tela = 0
contador_geral = 1;

horainicial=datetime.datetime.now()

"""  Comente esse treco do while ate o final caso queira testar um único link do imóvel """

registro = 0   
totapagina=0
print(f"HORA INICIAL {datetime.datetime.now()}")

logging.basicConfig(filename=caminho_da_planilha+'\\zap_log_error.log', level=logging.ERROR ,format='%(asctime)s - %(levelname)s - %(message)s')

if  tipo_processo == "gerarlinks" :
    
    try:
        # Criando o arquivo de texto para armazenar os links
        with open(caminho_da_planilha+'\\cidade_'+cidade_para_nome_arquivo+'.txt', 'w', encoding='utf-8') as arquivo_links:
            while pagina_atual >= 1:
                
                service = Service(chrome_driver_path)
                driver = webdriver.Chrome(service=service, options=chrome_options,keep_alive=True)#  
                # Montando a URL com o número da página atual
                url= caminho_do_filtro + str(pagina_atual)
                # O resto do seu código (dentro do loop) continua exatamente igual:
                driver.maximize_window()

                driver.get(url)
                time.sleep(random.uniform(1, 3))
                print("COLENTANDO OS LINK DOS IMOvES..")

                print((f"PAGINA ATUAL: {pagina_atual} "))  
        
                imoveis = driver.find_elements(By.XPATH, '//*[@id="__next"]/main/section/div/form/div[2]/div[4]/div[1]/div/div')  
            
                # Verifica se já coletou todos os imóveis da página atual
                if len(imoveis) <= 1:
                    print("Final de leitura dos registros.")
                    break 
                # Pagina os registro em cadas pagina ate 106 no maximo
                a = 0
                
                contreg=0;
                # Loop para iterar pelos elementos de imóvel
                for cont in range(1, 150):  # Itera até o final da lista de imóveis
                    try:
                        
                        # Coleta o link e outros dado        
                        a += 1
                                
                        driver.implicitly_wait(0)
                        link_element = driver.find_elements(By.XPATH, f'//*[@id="__next"]/main/section/div/form/div[2]/div[4]/div[1]/div/div[{cont}]/div/a')
                        
                        botao_menssagen = driver.find_elements(By.XPATH, f'//*[@id="__next"]/main/section/div/form/div[2]/div[4]/div[1]/div/div[{cont}]/div/div/div[2]/div[3]/div[2]/div[1]/button[2]')  
                        
                        botao_anuncio = driver.find_elements(By.XPATH, f'//*[@id="__next"]/main/section/div/form/div[2]/div[4]/div[1]/div/div[{cont}]/div/div/div/div/div[2]/div[3]/div[2]/button')   
                    
                        if botao_anuncio and len(botao_anuncio) > 0: 
                            button1 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="__next"]/main/section/div/form/div[2]/div[4]/div[1]/div/div[{cont}]/div/div/div/div/div[2]/div[3]/div[2]/button')))
                            texto = button1.text
                            
                            # Clica no botão de anúncio
                            try:
                                driver.implicitly_wait(4)  
                                if texto=="Exibir Anúncios": 
                                    driver.execute_script("arguments[0].scrollIntoView();", button1) 
                                    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE) # 
                                    button1.click()
                                    link_imovel = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,    '//*[@id="__next"]/main/section/aside/form/section/div/section[2]/section/div[5]/div[1]/a')))# Ajuste o tempo de espera conforme necessário
                                #
            
                            except Exception  as e:
                                print(f"Erro - {e}") 
                                
                            if link_imovel:
                               link_result = link_imovel.get_attribute('href')
                               print("IMOVEL COM ANÚNCIO")
                               print(f"{cont} - {link_result}")
                               contreg+=1
                               links_imoveis.add(link_result)
                               driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE) # Fechando o anúncio
                                
                                # Adicionando o link ao arquivo de texto
                               arquivo_links.write(link_result + '\n')
                                
                        # Coleta o link do imóvel (se não houver anúncio)
                        if link_element and len(link_element) > 0:
                           link = link_element[0].get_attribute('href')
                           links_imoveis.add(link)
                            
                           print(f"LINK: {cont} - {link}")
                           arquivo_links.write(link + '\n')
                           contreg+=1
                        
                        if a == 15:
                           driver.execute_script("arguments[0].scrollIntoView();", imoveis[-1])
                           time.sleep(2) 
                           a = 0    
                            
                        if cont ==105:
                           totapagina=totapagina+contreg;
                           contreg=0;
                           break# Não use 'continue' aqui, pois ele está fora do bloco 'try'   
                    
                    except Exception as e:  # Captura qualquer exceção
                        print(f"ERRO: {e} - {link}")
                        continue  # Continua para a próxima iteração do loop
                    
                    # Continua para o próximo registro
                print(f"TOTAL DE REGISTROS NAS PAGINAS: {totapagina}")
                pagina_atual+=1
            
    except Exception as e:
        print(f"ERRO GERAL: {e}")
        logging.error(f"Erro ao coletar link na geracao do linl: {e} - {link}")
        driver.quit()
    finally:
        print("Finalizando a coleta de dados.")
        print(f"Total de imóveis coletados: {contador_geral}")
        driver.quit()                


print("HORA INCIAL DE LEITURA: ", horainicial.strftime("%H:%M:%S"))
horaatual=datetime.datetime.now()
print("HORA FINAL DE LEITURA: ", horaatual.strftime("%H:%M:%S"))  


"""-----------------------------------------------------------------------------------"""
"""  Aqui e a segunda parte onde sera processado cada link para ser resgtado os dados """
"""-----------------------------------------------------------------------------------"""
# Mostra os registros catalogados no  link_imoveis
if tipo_processo == 'lerlinks':
    
    print("CRIANDO A PLANILHDA DO EXCEL.") 
    print(f"TOTAL DE REGISTRS COLETADOS: {len(links_imoveis)}")


    """ Array que contem os tipso de movel """
    tipos_imoveis = [
        "Apartamento",
        "Studio",
        "Kitnet",
        "Casa",
        "Sobrado",
        "Casa de Condomínio",
        "Casa de Vila",
        "Cobertura",
        "Flat",
        "Loft",
        "Terreno / Lote / Condomínio",
        "Fazenda / Sítio / Chácara",
        "Loja / Salão / Ponto Comercial",
        "Conjunto Comercial / Sala",
        "Casa Comercial",
        "Hotel / Motel / Pousada",
        "Andar / Laje Corporativa",
        "Andar / Laje corporativa",
        "Prédio Inteiro",
        "Terrenos / Lotes Comerciais",
        "Terreno / Lote Comercial",
        "Galpão / Depósito / Armazém",
        "Garagem"
    ]

    # Criando uma nova planilha Excel e definindo os cabeçalhos
    #Seta o diretorio onde será salvao a planilha
    output_dir = caminho_da_planilha
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Imóveis"
    cabecalhos = ["data_inclusao", "tipo_imovel", "cep_endereco", "logradouro_endereco", "numero_endereço", "bairro_endereco", "uf_endereco","cidade_endereco","dormitorios",
                "suites","vagas","area_util","vlr_venda","vlr_aluguel","vlr_condominio","vlr_iptu","anunciante","link","codigo_imovel"]
    sheet.append(cabecalhos)  # Adiciona os cabeçalhos à planilha

    """ Isso e para teste de um link especifico descomente aqui """
    #links_imoveis.add("https://www.zapimoveis.com.br/imovel/venda-casa-3-quartos-pantanal-miguel-pereira-130m2-id-2735351938/")
    # Coletando detalhes de cada imóvel
    registro_atual = 0
    cont_link=0

    # Limite máximo de threads simultâneas
    MAX_THREADS = 10 
    semaphore = Semaphore(MAX_THREADS)  # Define um semáforo com o limite de threads

    # Variável compartilhada para o contador de links
    cont_link = 0
    cont_link_lock = Lock()  # Cria uma trava para proteger a variável

    """-----------------------------------------------"""
    """ Funçao que traz os detalhes e esta nas theads """
    """-----------------------------------------------"""
    def processar_imovel(link,tipo_de_filtro):
        global registro_atual, cont_link
        area_total=""
        num_dormitorios=""
        num_suites=""
        num_vagas=""
        tipo_movel=""
        cepencontado=""
        preco=""
        data_cadastro = "" 
        logradouro = "" # Inicializa data_cadastro como None
        numero=""
        try:
            semaphore.acquire()
            service1 = Service(chrome_driver_path)
            driver1 = webdriver.Chrome(service=service1, options=chrome_options)
                 
            #time.sleep(1) 
            # Incrementa o contador de links com sincronização
            with cont_link_lock:  # Utiliza a trava para proteger o contador
                cont_link += 1
                print(f"LINK: {cont_link} - {link}") 
                
            driver1.get(link)  # Acessa a página do imóvel
            
            # ... (o restante do código para coletar dados do imóvel) ...
            items_lista =  WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]'))                                                )
            if items_lista is not None:
               driver1.implicitly_wait(0) 
               
               try:
                 html_content = items_lista.get_attribute('outerHTML')
               except Exception as e:
                 logging.error(f"No componente da tela:  ---> {link}   -   {e} ")    
                     
          
            
            try:
               precos_imovel = WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/div[3]/div[1]')))
               preco= separar_precos(precos_imovel.text)
                   
            except Exception:
               preco=""
              
        
            soup = BeautifulSoup(html_content, 'html.parser')       
            amenity_items = soup.find_all('p', class_='amenities-item')
    
            amenities_data = {}
        
            for item in amenity_items:
                property_name = item.get('itemprop')
        
            if property_name=='floorSize':
                property_value = item.find('span', class_='amenities-item-text').text.strip()    
                area_total = extrair_numeros(property_value)   
        
            if property_name=='numberOfRooms':
                property_value = item.find('span', class_='amenities-item-text').text.strip()        
                num_dormitorios = extrair_numeros(property_value) 

            if property_name=='numberOfSuites':
                property_value = item.find('span', class_='amenities-item-text').text.strip()        
                num_suites = extrair_numeros(property_value) 
                
            if property_name=='numberOfParkingSpaces':
                property_value = item.find('span', class_='amenities-item-text').text.strip()        
                num_vagas = extrair_numeros(property_value)           
    
            title = WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[2]/section/div[1]/h1')))
            address = WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/div[5]/div[1]/p')))
        
            tipo_valor_venda=WebDriverWait(driver1, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="business-type-info"]')))   
            if tipo_valor_venda.text == "Venda":
                valor_venda = WebDriverWait(driver1, 10).until(EC.presence_of_element_located((By.XPATH, '//p[@data-testid="price-info-value"]')))    
                valor_venda_limpo=valor_venda.text;
                valor_aluguel_limpo=""
                if valor_venda.text != "Sob consulta":
            
                    condo_fee = WebDriverWait(driver1,10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="condo-fee-price"]')))
                    if not condo_fee.text:
                       condo_fee = "Isento"  # Assinala um valor padrão se vazio
                    
                    iptu = WebDriverWait(driver1, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="iptu-price"]')))
                    if not iptu.text:
                       iptu = "Isento"  # Assinala um valor padrão se vazi  
                
                    if isinstance(condo_fee, str):  # Verifica se condo_fee é uma string
                        condo_fee = "Não foi possível recuperar"  # Assinala um valor padrão se vazio
                    else:
                        condo_fee = condo_fee.text  # Obtém o texto se for um WebElement
                
                    #if isinstance(iptu, str):  # Verifica se iptu é uma string
                     #   iptu = "Não foi possível recuperar"  # Assinala um valor padrão se vazio
                    #else:
                    #    iptu = iptu.text 
                        
            elif tipo_valor_venda.text == "Aluguel":
                                                                                                        
                valor_aluguel = WebDriverWait(driver1, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/div[3]/div/div[1]/div[1]/p[2]')))
                valor_aluguel_limpo=valor_aluguel.text
                valor_venda_limpo=""
                
                condo_fee = WebDriverWait(driver1, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="condo-fee-price"]')))
                if not condo_fee.text:
                   condo_fee = "Isento"  # Assinala um valor padrão se vazio
                
                iptu = WebDriverWait(driver1, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="iptu-price"]')))
                if not iptu.text:
                   iptu = "Isento"  # Assinala um valor padrão se vazio
                
                if isinstance(condo_fee, str):  # Verifica se condo_fee é uma string
                   condo_fee = "Não foi possível recuperar"  # Assinala um valor padrão se vazio
                else:
                   condo_fee = condo_fee.text  # Obtém o texto se for um WebElement
                
                if isinstance(iptu, str):  # Verifica se iptu é uma string
                   iptu = "Não foi possível recuperar"  # Assinala um valor padrão se vazio
                else:
                   iptu = iptu.text   
            
                area_valores=   WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]')))
                if not area_valores.text:
                    print("Nao existe")
            
            try:        
              data_extenso = WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[2]/section/div[4]/span[2]')))  
            except Exception:
               data_extenso=""
            
            try:
               data_cadastro = extrair_e_formatar_data(data_extenso.text)    
            except Exception:
               data_cadastro=""
            
            driver1.implicitly_wait(3)
            
            button2 = WebDriverWait(driver1,30).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[2]/section/div[3]/button')))
            if button2:
               button2.click()  
               time.sleep(3)
            
               anunciante_do_imovel = WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div/section/section/div[1]/div/p[1]')))
               zap_do_anunciante = WebDriverWait(driver1, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div/section/section/div[2]/p[3]')))
               logradouro, numero, bairro, cidade, estado = parse_address(address.text)
            
            #Trata o cep
            try:   
                cep = get_ceps_por_logradouro(logradouro, numero, cidade, estado)
            
                if cep:
                    cepencontado= cep['cep']
                else:
                    cepencontado="Nao encontrado"
            except Exception:
                cepencontado="Nao encontrado"
            
            try:   
               tipo_movel=encontrar_tipo_imovel(title.text,tipos_imoveis)
            except Exception:
               tipo_movel="inexistente"
            
            try:    
               parte_zap = remover_parte_texto(zap_do_anunciante.text, "No Zap: ")
            except Exception:
               parte_zap="inexistemte"    
        
            time.sleep(2) 
            if tipo_de_filtro=="venda":       
                precovenda=preco['Venda']  
                sheet.append([data_cadastro,tipo_movel,cepencontado, logradouro, numero,bairro,estado,cidade, num_dormitorios ,num_suites ,num_vagas ,area_total,precovenda,"", condo_fee, iptu,anunciante_do_imovel.text,link, parte_zap])
            
            if  tipo_de_filtro=="aluguel":
                precoaluguel=preco['Aluguel']  
                sheet.append([data_cadastro,tipo_movel,cepencontado, logradouro, numero,bairro,estado,cidade, num_dormitorios ,num_suites ,num_vagas ,area_total,"",precoaluguel, condo_fee, iptu,anunciante_do_imovel.text,link, parte_zap])   
                
            registro_atual += 1
            
            if registro_atual % 50 == 0:
                os.makedirs(output_dir, exist_ok=True)  # Cria o diretório se não existir
                workbook.save(os.path.join(output_dir, "listadeimoveis.xlsx")) 
                print(f"## PLANILHA SALVA COM REGISTROS (Registro {registro_atual}) ##")

            semaphore.release()  # Libera um espaço no semáforo após terminar
        except Exception as e:
       
            if tipo_de_filtro=="venda":       
                precovenda=preco['Venda']  
                sheet.append([data_cadastro,tipo_movel,cepencontado, logradouro, numero,bairro,estado,cidade, num_dormitorios ,num_suites ,num_vagas ,area_total,precovenda,"", condo_fee, iptu.text,anunciante_do_imovel.text,link, parte_zap])
            
            if  tipo_de_filtro=="aluguel":
                precoaluguel=preco['Aluguel']  
                sheet.append([data_cadastro,tipo_movel,cepencontado, logradouro, numero,bairro,estado,cidade, num_dormitorios ,num_suites ,num_vagas ,area_total,"",precoaluguel, condo_fee, iptu.text, anunciante_do_imovel.text,link, parte_zap])           
            #driver1.quit()
            logging.error(f"Na coleta detalhes do imóvel dados faltando mas foi adicionado na planilha:  ---> {link}   -   {e} ")

            semaphore.release()  
        #finally:
        #    semaphore.release()# Libera um espaço no semáforo após terminar
    # Cria threads para processar cada link
    # Cria threads para processar cada link em lotes de 5
    
    # Lê os links do arquivo texto
    links_imoveis = [] 
    with open(caminho_da_planilha + '\\cidade_'+cidade_para_nome_arquivo+'.txt', 'r', encoding='utf-8') as arquivo_links:
        for link in arquivo_links:
            links_imoveis.append(link.strip()) 

    lote_atual = 0
    tamanho_lote =int(maximo_theads)  # Ajuste o tamanho do lote
    while lote_atual * tamanho_lote < len(links_imoveis):
        try:
            print(f"Iniciando lote {lote_atual}")
            lote = links_imoveis[lote_atual * tamanho_lote: (lote_atual + 1) * tamanho_lote]
            threads = []
            for link in lote:
                # Tempo de espera aleatório
                time.sleep(random.uniform(2, 5))  # Espere entre 2 e 5 segundos

                thread = Thread(target=processar_imovel, args=(link, tipo_de_filtro))
                threads.append(thread)
                thread.start()

            # Aguarda todas as threads do lote terminarem
            for thread in threads:
                thread.join(timeout=60)
                if thread.is_alive():
                    print(f"A thread para o link {link} não terminou em 60 segundos.")

            lote_atual += 1  # Incrementa o lote atual após o processamento do lote
        except Exception as e:
            #print(f"ERROR: {e}")
            #traceback.print_exc()
            logging.error(f"Erro ao coletar link: {e}")
        finally:
            print(f"Terminando lote {lote_atual}")
    # Salva a planilha final
    os.makedirs(output_dir, exist_ok=True)  # Create the directory if it doesn't exist
    workbook.save(os.path.join(output_dir, "listadeimoveis.xlsx"))
    print("Planilha Excel criada com sucesso!")

print("FINALIZADO COM SUCESSO..")
print("HORA INCIAL DE LEITURA: ", horainicial.strftime("%H:%M:%S"))
horafinal=datetime.datetime.now()
print("HORA FINAL DE LEITURA: ", horafinal.strftime("%H:%M:%S")) 

time.sleep(10)
sys.exit()